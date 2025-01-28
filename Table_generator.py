from enum import Enum
from openpyxl import load_workbook
from utils.convertUserName import osuWikiUser

def convert_align_text(align_text:str):
    if align_text is None:
        return tableAlign.left
    elif align_text == 'left':
        return tableAlign.left
    elif align_text == 'center':
        return tableAlign.center
    elif align_text == 'right':
        return tableAlign.right
    else:
        return tableAlign.left

class tableAlign(Enum):
    left = 0,
    center = 1,
    right = 2

class wikiTable:
    columns = []

class wikiColumn:
    head = ''
    align = tableAlign.left
    bold = False
    user_column = False
    using_UID = False

    def __init__(self):
        self.body = []

    def __str__(self):
        return f'{self.head}|{self.body}'

    def __getitem__(self, key:int):
        return self.body[key]

    def __len__(self):
        return len(self.body)

    def __add__(self, other:str):
        self.body.append(other)
        return self

    def get_head_str(self,append = False):
        result = '' if append else '|'
        return result + ' ' + self.head + ' |'

    def get_align_str(self,append = False):
        result = '' if append else '|'
        if self.align == tableAlign.right:
            result += ' --: |'
        elif self.align == tableAlign.center:
            result += ' :-: |'
        else:
            result += ' :-- |'
        return result

    def get_body_str(self,append = False, key = 0):
        result = '' if append else '|'
        if key >= len(self):
            content = ''
        else:
            content = self.cover_user_string(self[key],is_uid=self.using_UID) if self.user_column else self.body[key]
        if len(content) and self.bold:
            content = f'**{content}**'
        return result + ' ' + content + ' |'

    def cover_user_string(self,user_string,split_char = ",",is_uid = False):
        if len(user_string) == 0:
            return '*TBA*'
        else:
            split_user_string = user_string.split(split_char)
            split_user_string.sort(key=lambda x: x.lower())
            result = ''
            for i in range(0,len(split_user_string)):
                try:
                    if (is_uid):
                        user = osuWikiUser(split_user_string[i],is_uid=True)
                    else:
                        user = osuWikiUser(split_user_string[i])
                    print(f'get osu user "{user.uname}" success.')
                except ValueError as e:
                    print('unable to find player name' + user.uname + 'details below')
                    print(e)
                    continue
                result += str(user)
                if i < len(split_user_string) - 1:
                    result += ', '
        return result

    def set_column_str(self, string_array, body_rows = None):
        append = True
        if string_array is None or len(string_array) == 0:
            string_array = []
            append = False
            while len(string_array) < len(self) + 2:
                string_array.append('')
        if body_rows is None:
            body_rows = len(self)
        string_array[0] += self.get_head_str(append)
        string_array[1] += self.get_align_str(append)
        for i in range (0,body_rows):
            string_array[2+i] += self.get_body_str(append,i)
        return string_array

class wikiTable:
    columns = []
    max_body_row = 0

    def __add__(self,column:wikiColumn):
        self.columns.append(column)
        return self

    def __len__(self):
        return len(self.columns)

    def __getitem__(self, key:int):
        return self.columns[key]

    def __str__(self):
        str_array = self.output_table()
        result = ''
        for str_obj in str_array:
            result += str_obj + '\n'
        return result

    def load_table(self, sheet_dir):
        wb = load_workbook(filename=sheet_dir + '/wiki_table.xlsx', read_only=True)
        ws = wb['Data']
        m_row = ws.max_row
        if m_row is None:
            m_row = 500
        finish_reading = False
        column_index = 3
        while not finish_reading:
            current_column = wikiColumn()
            align_text = ws.cell(row=2, column=column_index).value
            current_column.align = convert_align_text(align_text)
            bold_text = ws.cell(row=3, column=column_index).value
            current_column.bold = True if bold_text == '√' else False
            user_column_text = ws.cell(row=4, column=column_index).value
            current_column.user_column = True if user_column_text == '√' else False
            using_UID_text = ws.cell(row=5, column=column_index).value
            current_column.using_UID = True if using_UID_text == '√' else False


            empty_head = False
            head_text = ws.cell(row=6, column=column_index).value
            if head_text is None:
                empty_head = True
                head_text = ''
            current_column.head = head_text

            empty_body = True
            for cur_row_index in range(7,m_row+1):
                body_text = ws.cell(row=cur_row_index, column=column_index).value
                if body_text is None:
                    body_text = ''
                else:
                    if not isinstance(body_text, str):
                        body_text = str(body_text)
                    empty_body = False
                    self.max_body_row = cur_row_index-6
                current_column += body_text

            if empty_head and empty_body:
                return self
            else:
                self.columns.append(current_column)
                column_index += 1
        return self

    def output_table(self):
        result = []
        if len(self) == 0:
            return result
        else:
            for column_index in range(0,len(self)):
                result = self[column_index].set_column_str(result,body_rows=self.max_body_row)
        return result

if __name__ == '__main__':
    sheet_dir = 'data'
    output_dir = 'output'

    table = wikiTable()
    table.load_table(sheet_dir=sheet_dir)
    output_strings = table.output_table()
    with open(output_dir+'/wiki_table.txt', "w+" , encoding='utf-8') as f:
        for output_string in output_strings:
            f.write(output_string + '\n')
