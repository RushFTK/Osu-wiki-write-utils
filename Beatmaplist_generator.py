from dis import UNKNOWN

from openpyxl import load_workbook
from utils.common import generate_osu_api
from utils.common import convert_char
from ossapi import GameMode

api = generate_osu_api()

class wikiBeatmaps:
    mode = GameMode.OSU
    sid = 0
    bid = 0
    artist = ''
    title = ''
    mapper = ''
    difficult_name = ''

    def __init__(self, bid = 0):
        if not (bid == 0):
            self.set_beatmap_by_id(bid)

    def __str__(self):
        display_name = f'{self.artist} - {self.title} ({self.mapper})[{self.difficult_name}]'
        url = f'https://osu.ppy.sh/beatmapsets/{self.sid}#{self.mode.value}/{self.bid}'
        return f'[{convert_char(display_name)}]({url})'

    def set_beatmap_by_id(self,bid):
        try:
            beatmap = api.beatmap(bid)
            beatmapset = api.beatmapset(beatmap_id=bid)
            print (f'load beatmap {bid} success')
            self.sid = beatmap.beatmapset_id
            self.bid = beatmap.id
            self.mode = beatmap.mode
            self.artist = beatmapset.artist
            self.title = beatmapset.title
            self.mapper = beatmapset.creator
            self.difficult_name = beatmap.version
            cleantags = ['[4K] ', '[7K] ']
            for cleantag in cleantags:
                self.difficult_name = self.difficult_name.replace(cleantag, '')
        except ValueError as e:
            print(e)
        return self

class wikiBeatmapLists:
    modlist = {
        'NM' : 'No Mod',
        'HD' : 'Hidden',
        'HR' : 'Hard Rock',
        'DT' : 'Double Time',
        'FM' : 'Free Mod',
        'MM' : 'Mixed Mod',
        'RC' : 'Rice',
        'HB' : 'Hybrid',
        'LN' : 'Long Note',
        'SV' : 'SV',
        'TB' : 'Tiebreaker'
    }

    def __init__(self):
        self.maplist = {}

    def __getattr__(self, item):
        return self.maplist[item]

    def get_modslot_by_id(self,id:str):
        result = ''
        for id_prefix in self.modlist.keys():
            if id.startswith(id_prefix):
                result = self.modlist[id_prefix]
                break
        return result

    def save_beatmap(self,bid,mod:str=None,id:str=None):
        if isinstance(bid,str):
            bid = int(bid, 10)
        betamap = wikiBeatmaps(bid)
        if mod is None or len(mod) == 0:
            if id is None or len(id) == 0:
                mod = ''
            else:
                mod = self.get_modslot_by_id(id)
        if not mod in self.maplist.keys():
            self.maplist[mod] = []
        self.maplist[mod].append(betamap)
        return True

    def output_mod(self,mod):
        result = ''
        if not mod in self.maplist.keys():
            print (f'mod "{mod}" isn\'t in map list!')
            return result
        if mod == '':
            beatmaps = self.maplist[mod]
            for i in range(0,len(beatmaps)):
                result += f'{i+1}. {beatmaps[i]}\n'
        else:
            result += f'- {mod}\n'
            beatmaps = self.maplist[mod]
            for i in range(0, len(beatmaps)):
                result += f'  {i+1}. {beatmaps[i]}\n'
        return result

    def output_list(self):
        result = ''
        for mod in self.maplist.keys():
            result += self.output_mod(mod)
        return result

def load_maplist(dir):
    maplist = wikiBeatmapLists()
    wb = load_workbook(filename=sheet_dir + '/beatmap_list.xlsx', read_only=True)
    ws = wb['Data']
    m_row = ws.max_row
    if m_row is None:
        m_row = 500
    for i in range(3,m_row):
        bid = ws.cell(row=i, column=4).value
        if bid == None:
            break
        mod = ws.cell(row=i, column=2).value
        id = ws.cell(row=i, column=3).value
        maplist.save_beatmap(bid, mod,id)
    return maplist


if __name__ == '__main__':
    sheet_dir = 'data'
    output_dir = 'output'

    maplist = load_maplist(sheet_dir)
    output_string = maplist.output_list()
    with open(output_dir+'/beatmap_list.txt', "w+" , encoding='utf-8') as f:
        f.write(output_string)
